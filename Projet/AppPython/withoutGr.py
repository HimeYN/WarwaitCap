import sys
import os

from time import sleep
from tqdm import tqdm
from numpy import *

import pandas as pd
import openpyxl 

import mysql.connector as mysql

class withouGr():
    def __init__(self,path):
        ######## Database constants
        self.HOST = "localhost"       ## replace by the domain or the ip address

        self.DATABASE = ["cis_nord_warwait","skillmatrix"]  ## replace by the name of the database

        self.USER = "admin"          ## replace by a user administrator of the db

        self.PASSWORD = "passwd"  ## replace by the password tied with the user

        self.path = path


    def fill_database_matrix(self):
        try:
            database = mysql.connect(host=self.HOST,
                                    database=self.DATABASE[1],
                                    user=self.USER,
                                    password=self.PASSWORD)
            print("Connected to database... OK")
        except mysql.Error as e:
            raise Exception("Cant connect to the database : {}".format(e))
        cursor = database.cursor()
        db = pd.read_excel(self.path, sheet_name="Matrice",skiprows=1,nrows=115)
        sf = pd.read_excel(self.path,sheet_name='Matrice',skiprows=1,nrows=115)
        tables = []
        for names in sf.to_numpy():
            name = names[0]
            if type(name) == str:
                if char.isupper(name):
                    tables.append(name)
        db_list = db.to_numpy()
        names = db.columns.to_numpy()

        for i,n in enumerate(names):
            names[i] = n.replace("\u200b","")

        db_index=0
        all_tec = []
        all_values = []

        for table_index in range(len(tables)-1):
            techno = []
            values = []
            while table_index+1<len(tables) and db_list[db_index][0] != tables[table_index+1]:
                

                techno.append(db_list[db_index][0]) ## add the techno in the list
                all_tec.append(db_list[db_index][0])
                values.append(db_list[db_index])
                all_values.append(db_list[db_index])
                # print(db_list[db_index])
                # for i,elements in enumerate(db_list[db_index]):
                #     if i == 0:
                    ##print(elements)


                db_index+=1
            # print(tables[table_index])
            # print(techno)
            
            if table_index+1<len(tables):
                params = "\'{}\'"
                sql = "CREATE TABLE IF NOT EXISTS "+ tables[table_index].replace(" ","_").replace("È","E").replace("&","ET").replace("É","E") +" (name TEXT"
                insert_sql = "INSERT INTO "+tables[table_index].replace(" ","_").replace("È","E").replace("&","ET").replace("É","E")+"( name "
                for tec in techno:
                    if tec == "DISPONIBILITÉ":
                        sql+=", "+tec.replace(',','').replace('#','sharp').replace('++','pp').replace(' ', '_').replace('.','_').replace('/','_').replace('\\','_').replace('-','_').replace('__','_').replace('__','_').replace("\xc2\xa0","").replace('(','').replace(')','').lower()+" TEXT"
                    else:
                        sql+=", "+tec.replace(',','').replace('#','sharp').replace('++','pp').replace(' ', '_').replace('.','_').replace('/','_').replace('\\','_').replace('-','_').replace('__','_').replace('__','_').replace("\xc2\xa0","").replace('(','').replace(')','').lower()+" INTEGER"
                    insert_sql+=", "+tec.replace(',','').replace('#','sharp').replace('++','pp').replace(' ', '_').replace('.','_').replace('/','_').replace('\\','_').replace('-','_').replace('__','_').replace('__','_').replace("\xc2\xa0","").replace('(','').replace(')','').lower()
                    if tec == "DISPONIBILITÉ" :
                        params+=", \'{}\'"
                    else:
                        params+=", {}"
                sql +=")"
                insert_sql+=") VALUES (" + params + ")"
                #print(sql)

                
            

            
            #print(names)
            #print(values)
            cursor.execute("DROP TABLE IF EXISTS "+tables[table_index].replace(" ","_").replace("È","E").replace("&","ET").replace("É","E"))
            cursor.execute(sql)
            for name in range(len(names)-1):
                tp_res = ()
                tp_res+=(names[name+1],)
                for v in values:
                    if v[0] == 'DISPONIBILITÉ':
                        tp_res+= (" ",) if str(v[name+1])=="nan" else (str(v[name+1]).replace(" ",""),)
                    else:
                        if type(v[name+1]) != str:
                            tp_res+= (int(0),) if isnan(v[name+1]) else (int(v[name+1]),) 
                        else:
                            tp_res+= ("",) if str(v[name+1])=="nan" else (str(v[name+1]),)
                
                #print(self.fill_query(insert_sql,tp_res))
                cursor.execute(self.fill_query(insert_sql,tp_res))            

            db_index+=1

        cursor.execute("DROP TABLE IF EXISTS matrix_tech")
        sql = "CREATE TABLE IF NOT EXISTS matrix_tech (name TEXT"
        for elt in tables:
            if elt == 'ANGLAIS' or elt == 'DISPONIBILITÉ':
                sql+=", "+elt.replace(" ","_").replace("È","E").replace("&","ET").replace("É","E")+" TEXT"
            else :
                sql+=", "+elt.replace(" ","_").replace("È","E").replace("&","ET").replace("É","E")+" FLOAT"
        sql +=')'

        cursor.execute(sql)
        all_vals=[]
        for t in range(len(tables)-1):
            tables_val = []
            for name in range(len(names)-1):
                #print('SELECT * from '+t+' WHERE name ='+names[name+1]+'')
                cursor.execute('SELECT * from '+tables[t].replace(" ","_").replace("È","E").replace("&","ET").replace("É","E")+' WHERE name =\''+names[name+1]+'\'')
                tables_val.append(cursor.fetchall())
            
            all_vals.append(tables_val)

        #print(all_vals)
        for name in range(len(names)-1):
            sql = "INSERT INTO matrix_tech (name"
            values="'{}'"
            tp_res=()
            tp_res+=(names[name+1],)
            for t in range(len(tables)-1):
                sql+=", "+tables[t].replace(" ","_").replace("È","E").replace("&","ET").replace("É","E")
                if tables[t] == 'ANGLAIS' or tables[t] == 'DISPONIBILITÉ':
                    values+=", '{}'"
                    if  tables[t] == 'ANGLAIS':
                        tp_res+=(db_list[len(db_list-2)],) if str(db_list[len(db_list-2)]) != "nan" else (" ",)
                    else:
                        tp_res+=(all_vals[t][name][0][1],)
                else :
                    values+=", {}"
                    res=0
                    #print(all_vals[t][name][0])
                    for i in range(1,len(all_vals[t][name][0])):
                        #print(all_vals[t][name][0][i])
                        res+=all_vals[t][name][0][i]
                        
                    tp_res+=(round(res/(len(all_vals[t][name][0])-1),2),)
            tp_res+=(db_list[len(db_list)-2][name+1],) if str(db_list[len(db_list)-2][name+1]) != "nan" else (" ",)
            sql += ", ANGLAIS) VALUES ( "+values+", '{}')"
            #print(self.fill_query(sql,tp_res))
            #print(db_list[113])
            cursor.execute(self.fill_query(sql,tp_res))

        _ =all_tec.pop(0)
        #print(all_tec)

        cursor.execute("DROP TABLE IF EXISTS skillmatrix")
        create_table_skillmatrix="CREATE TABLE IF NOT EXISTS skillmatrix (name TEXT"
        for t in all_tec:
            create_table_skillmatrix+=", "+t.replace(',','').replace('#','sharp').replace('++','pp').replace(' ', '_').replace('.','_').replace('/','_').replace('\\','_').replace('-','_').replace('__','_').replace('__','_').replace("\xc2\xa0","").replace('(','').replace(')','').lower()+" INTEGER"

        cursor.execute(create_table_skillmatrix+")")
        #print(create_table_skillmatrix+")")

        for name in range(len(names)-1):
            sql = "INSERT INTO skillmatrix (name"
            params="'{}'"
            tp_res=(names[name+1],)
            for i in range(len(all_tec)):
                sql+=", "+all_tec[i].replace(',','').replace('#','sharp').replace('++','pp').replace(' ', '_').replace('.','_').replace('/','_').replace('\\','_').replace('-','_').replace('__','_').replace('__','_').replace("\xc2\xa0","").replace('(','').replace(')','').lower()
                params+=", {}"
                if type(all_values[i+1][name+1]) != str:
                    tp_res+= (int(0),) if isnan(all_values[i+1][name+1]) else (int(all_values[i+1][name+1]),) 
                else:
                    tp_res+= (0,) if str(all_values[i+1][name+1])=="nan" else (int(all_values[i+1][name+1]),)
            #print(tp_res)
            #print(len(tp_res),len(all_tec))

            sql+=") VALUES ("+params+")"
            cursor.execute(self.fill_query(sql,tp_res))


        database.commit()
        print('database has been added')

        if database.is_connected():
                    cursor.close()
                    database.close()
                    print("MySQL connection is closed")




    def fill_database_warwait(self):

        try:
            database = mysql.connect(host=self.HOST,
                                    database=self.DATABASE[0],
                                    user=self.USER,
                                    password=self.PASSWORD)
            print("Connected to database... OK")
        except mysql.Error as e:
            raise Exception("Cant connect to the database : {}".format(e))
        cursor = database.cursor()
        cursor.execute("DROP TABLE IF EXISTS warwait")  # Delete former table
        cursor.execute("DROP TABLE IF EXISTS status")
        cursor.execute("CREATE TABLE IF NOT EXISTS status (state TEXT)")
        cursor.execute("INSERT INTO status (state) VALUES (%s)",('indispo',))
        cursor.execute("INSERT INTO status (state) VALUES (%s)",('mission',))
        cursor.execute("INSERT INTO status (state) VALUES (%s)",('IP+ indispo',))
        cursor.execute("INSERT INTO status (state) VALUES (%s)",('IP+',))
        nbr = pd.ExcelFile(self.path).sheet_names
        name_in_db= []
        for _index_sheet_ in tqdm(range(len(nbr)-1)):
            sleep(0.1)

            try:
                week = pd.read_excel(self.path, sheet_name=_index_sheet_, skiprows=8, nrows=1)

                db = pd.read_excel(self.path, sheet_name=_index_sheet_, skiprows=9)

                xl = pd.ExcelFile(self.path).sheet_names[_index_sheet_]
                wb = openpyxl.load_workbook(self.path)

                licol = {}

                fs = wb[xl]
                fs_count_row = fs.max_row
                fs_count_col = fs.max_column
                for row in range(0,fs_count_row):
                    for column in range(0,fs_count_col):
                        cell_color = fs.cell(column=column+1, row=row+1)
                        bgColor = cell_color.fill.bgColor.index
                        fgColor = cell_color.fill.fgColor.index
                        if (bgColor=='00000000') or (fgColor=='00000000'):
                            continue
                        else:
                            #print("Background color index of cell (",row,column, ") is", fgColor)
                            if fgColor == 'FFFFFF00':
                                licol[(row,column)] = 'indispo'
                            elif fgColor == 'FFCCFFCC':
                                licol[(row,column)]='mission'

#                        print(licol)

            except PermissionError as e:
                raise PermissionError("Cant open the file, it is already running.")
    


            lis = db.to_numpy()
            week_list = week.to_numpy()

            create_table_warwait = """ CREATE TABLE IF NOT EXISTS warwait (nom TEXT, grade TEXT, site TEXT,s1 TEXT,s2 TEXT,s3 TEXT,s4 TEXT,s5 TEXT,s6 TEXT,s7 TEXT,s8 TEXT,s9 TEXT,s10 TEXT,s11 TEXT,s12 TEXT,s13 TEXT,s14 TEXT,s15 TEXT,s16 TEXT,s17 TEXT,s18 TEXT,s19 TEXT,s20 TEXT,s21 TEXT,s22 TEXT,s23 TEXT,s24 TEXT,s25 TEXT,s26 TEXT,s27 TEXT,s28 TEXT,s29 TEXT,s30 TEXT,s31 TEXT,s32 TEXT,s33 TEXT,s34 TEXT,s35 TEXT,s36 TEXT,s37 TEXT,s38 TEXT,s39 TEXT,s40 TEXT,s41 TEXT,s42 TEXT,s43 TEXT,s44 TEXT,s45 TEXT,s46 TEXT,s47 TEXT,s48 TEXT,s49 TEXT,s50 TEXT,s51 TEXT,s52 TEXT,reussite FLOAT,positionnement TEXT,competences TEXT,cv_code INTEGER,pe TEXT,en_mission TEXT,afficher TEXT, id integer auto_increment primary key) """
            try:
                cursor.execute(create_table_warwait)                  

            except mysql.Error as e:  # printing reason of error if it happen
                raise Exception("Cant create the table : {}".format(e))

            database.commit() 
            
            # taking weeks number on warwait and throwing away all empty cells in the list
            week_list = week_list[0]
            for loop in week_list:
                if type(loop) == str:
                    index = argwhere(week_list==loop)
                    week_list= delete(week_list, index)

            week_list = week_list[~pd.isnull(week_list)]
            # transform float to int
            week_list = week_list.astype(int)
            row = 9
            for ligne in lis:
                row +=1
                ## getting all informations from the line
                ## Ugly code here, need to find a better and more elegant way
                name = ligne[0]
                if name == "TOTAL Inters Nord et Est":
                    break

                

                grade = "NULL" if str(ligne[1])=="nan" else ligne[1]
                site = "NULL" if str(ligne[2])=="nan" else ligne[2]
                weeks= ()
                only_weeks = ()
                for index_w in range(len(week_list)):
                    #print(row, 3+index_w)
                    if (row, index_w+3) in licol:
                        if str(ligne[index_w+3]) == "nan":
                            only_weeks += (licol[(row, index_w+3)],)
                            #print('ici ',licol[(row, index_w)])

                        else :
                            only_weeks += (str(ligne[index_w+3])+" "+licol[(row, index_w+3)],)
                    else:
                        only_weeks += (str(ligne[index_w+3]),)

                offset = len(week_list)+3 
                success_rate = 0. if str(ligne[offset])=="nan" else ligne[offset]
                company = "NULL" if str(ligne[offset+1])=="nan" else ligne[offset+1]
                skills = "NULL" if str(ligne[offset+2])=="nan" else ligne[offset+2]
                cv_code = "34"
                pe = "" if str(ligne[offset+4]) == "nan" else str(ligne[offset+4])

                # generating weeks => Null if there're no info in warwait
                for _ in range(1, min(week_list)):
                    weeks = weeks + ("NULL",)

                weeks += only_weeks
                for i,w in enumerate(weeks):
                    if w == "nan":
                        week[i] = "NULL"

                for i,w in enumerate(only_weeks):
                    if w == "nan":
                        week[i] = "NULL"

                for _ in range(len(weeks)+1,53): ## range is lower <= x < upper
                    weeks = weeks + ("NULL",)
                params = (name,grade,site) + weeks + (success_rate,company,skills,cv_code,pe,"false",'true')


                ## Insert values into database
                if _index_sheet_ == 0:   
                    insert_warwait = "INSERT INTO warwait (nom, grade, site,s1,s2,s3,s4,s5,s6,s7,s8,s9,s10,s11,s12,s13,s14,s15,s16,s17,s18,s19,s20,s21,s22,s23,s24,s25,s26,s27,s28,s29,s30,s31,s32,s33,s34,s35,s36,s37,s38,s39,s40,s41,s42,s43,s44,s45,s46,s47,s48,s49,s50,s51,s52,reussite,positionnement,competences,cv_code,pe,en_mission,afficher) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    name_in_db.append(name)
                else:
                    if not (name in name_in_db) :
                        insert_warwait = "INSERT INTO warwait (nom, grade, site,s1,s2,s3,s4,s5,s6,s7,s8,s9,s10,s11,s12,s13,s14,s15,s16,s17,s18,s19,s20,s21,s22,s23,s24,s25,s26,s27,s28,s29,s30,s31,s32,s33,s34,s35,s36,s37,s38,s39,s40,s41,s42,s43,s44,s45,s46,s47,s48,s49,s50,s51,s52,reussite,positionnement,competences,cv_code,pe,en_mission,afficher) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                        name_in_db.append(name)
                    else :
                        w_str=""
                        for count in week_list:
                            w_str+="s"+str(count) +" =%s, "


                        insert_warwait = "UPDATE warwait SET grade = %s, site = %s, "+w_str+"reussite=%s, positionnement=%s, competences=%s, cv_code=%s, pe=%s, en_mission=%s, afficher=%s WHERE nom = %s"

                        params = (grade,site) + only_weeks + (success_rate,company,skills,str(34),pe,"false",'true',name)
                try:
                    cursor.execute(insert_warwait,params)
                    

                except mysql.Error as e:
                    raise Exception("Cant insert in the table : {}".format(e))
                    
            
            database.commit()
        print('database has been added')
        if database.is_connected():
            cursor.close()
            database.close()
            print("MySQL connection is closed")

    def fill_query(self, src, tp):
        i = 0
        j = 0
        while i < len(src) and len(tp)>0:
            c = src[i]
            if c == "{":                
                j = i+len(str(tp[0]))+1
                src = src[:i] + str(tp[0]) + src[i+2:]
                tp = tp[1:]
                i=j
            else:
                i +=1
        return src
    

if __name__ == "__main__":
    nb_args = len(sys.argv)

    if nb_args < 3:
        print('use : python withoutGr.py -W|-MC \'path_to_excel\'')
    else :
        if sys.argv[1] == "-W":
            setup = withouGr(sys.argv[2])
            setup.fill_database_warwait()

        elif sys.argv[1] == "-MC":
            setup = withouGr(sys.argv[2])
            setup.fill_database_matrix()
        else:
            print('use : python withoutGr.py -W|-MC \'path_to_excel\'')
